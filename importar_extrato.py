"""
Importador de extrato bancário — Escrow Monitor / Liqi Digital Assets

Converte o arquivo "Extrato de Lançamentos" do Itaú Escrow Advanced (.xlsx)
em data/extrato.json, usado pelo escrow_monitor.py para conciliar os eventos
extraídos dos e-mails contra o que de fato foi lançado na conta.

Uso:
    python importar_extrato.py caminho/para/Extrato_Lancamentos_8541_835719_30-07-2026.xlsx
"""

import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

EXTRATO_FILE = Path(__file__).parent / "data" / "extrato.json"

# O extrato registra a transferência judicial em duas linhas no mesmo dia:
# um DESBLOQUEIO de estorno e a TRANSF propriamente dita. O desbloqueio de
# estorno não é liberação de recurso e precisa ser neutralizado, senão o total
# de desbloqueios fica inflado pelo valor de todas as transferências.
TIPOS = {
    "BLOQUEIO": "BLOQUEIO",
    "DESBLOQUEIO": "DESBLOQUEIO",
    "TRANSF": "TRANSFERENCIA",
}

# Lançamentos que não são constrição judicial, mas compõem o caixa da conta e
# por isso são necessários para aferir o Saldo Mínimo Retido (cl. 6.2/6.3 do
# Contrato de Cessão Fiduciária; cl. 2.1 do Anexo I do Contrato de Custódia).
TIPOS_CAIXA = {
    "RECEBIMENTOS": "CREDITO",
    "RENDIMENTOS": "RENDIMENTO",
}

# Saldo Mínimo Retido exigido na conta (cl. 6.3 do Contrato de Cessão
# Fiduciária de Direitos Creditórios) e o piso inicial da cl. 6.2.
SALDO_MINIMO_RETIDO = 30_000_000.00
SALDO_MINIMO_RETIDO_INICIAL = 23_000_000.00


def classificar(lancamento):
    """Mapeia a descrição do lançamento para um dos três tipos financeiros."""
    for prefixo, tipo in TIPOS.items():
        if lancamento.startswith(prefixo):
            return tipo
    return None


def classificar_caixa(lancamento):
    """Mapeia crédito de recebível e rendimento de aplicação automática."""
    for prefixo, tipo in TIPOS_CAIXA.items():
        if lancamento.startswith(prefixo):
            return tipo
    return None


def apurar_garantia(totais, fluxos, saldo_conta):
    """Afere o colchão de garantia contra o Saldo Mínimo Retido.

    A cláusula 2.9 do Anexo I do Contrato de Custódia (e a 1.4.1) determina que
    valores bloqueados por ordem judicial NÃO compõem o Saldo Mínimo. O colchão
    efetivo é, portanto, apenas o saldo livre em conta.

    O resíduo é uma prova de caixa: créditos + rendimentos - transferências
    deveria igualar saldo livre + saldo bloqueado. Sobra positiva relevante
    indica recursos fora da conta corrente (aplicação financeira).
    """
    creditos = round(sum(f["valor"] for f in fluxos if f["tipo"] == "CREDITO"), 2)
    rendimentos = round(sum(f["valor"] for f in fluxos if f["tipo"] == "RENDIMENTO"), 2)
    bloqueado = totais["SALDO_BLOQUEADO"]
    colchao = round(saldo_conta, 2)
    residuo = round(creditos + rendimentos - totais["TRANSFERENCIA"] - colchao - bloqueado, 2)
    return {
        "saldo_minimo_exigido": SALDO_MINIMO_RETIDO,
        "saldo_minimo_inicial": SALDO_MINIMO_RETIDO_INICIAL,
        "creditos": creditos,
        "rendimentos": rendimentos,
        "saldo_livre": colchao,
        "saldo_bloqueado": bloqueado,
        "colchao_efetivo": colchao,
        "deficit": round(max(0.0, SALDO_MINIMO_RETIDO - colchao), 2),
        "cobertura": round(colchao / SALDO_MINIMO_RETIDO, 6) if SALDO_MINIMO_RETIDO else 0.0,
        "fora_da_conta": residuo,
    }


def parse_data(valor):
    """Converte a data do extrato (dd/mm/aaaa ou datetime) para ISO."""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor).strip()
    return datetime.strptime(texto, "%d/%m/%Y").strftime("%Y-%m-%d")


def ler_metadados(ws):
    """Lê o cabeçalho do extrato (titular, agência, conta, período, geração)."""
    meta = {}
    rotulos = {
        "Atualiza": "gerado_em",
        "Nome": "titular",
        "Ag": "agencia",
        "Conta": "conta",
        "Periodo": "periodo",
        "Período": "periodo",
    }
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        rotulo = str(row[0] or "").strip().rstrip(":")
        for prefixo, chave in rotulos.items():
            if rotulo.startswith(prefixo) and chave not in meta:
                meta[chave] = str(row[1] or "").strip()
    return meta


def importar(caminho_xlsx):
    caminho = Path(caminho_xlsx)
    if not caminho.exists():
        print(f"ERRO: arquivo não encontrado: {caminho}")
        sys.exit(1)

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[wb.sheetnames[0]]
    meta = ler_metadados(ws)

    brutos = []
    fluxos = []
    saldo_conta = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        descricao = str(row[1] or "").strip()
        if not descricao:
            continue
        if descricao.startswith("SALDO EM CONTA CORRENTE"):
            saldo_conta = row[5]
            continue
        if row[4] is None:
            continue
        tipo = classificar(descricao)
        if tipo:
            brutos.append({
                "data": parse_data(row[0]),
                "tipo": tipo,
                # o extrato traz bloqueio/transferência com sinal negativo
                "valor": round(abs(float(row[4])), 2),
            })
            continue
        tipo_caixa = classificar_caixa(descricao)
        if tipo_caixa:
            fluxos.append({
                "data": parse_data(row[0]),
                "tipo": tipo_caixa,
                "valor": round(float(row[4]), 2),
                "descricao": descricao,
            })

    lancamentos, estornos = neutralizar_estornos(brutos)

    totais = {t: round(sum(l["valor"] for l in lancamentos if l["tipo"] == t), 2)
              for t in ("BLOQUEIO", "DESBLOQUEIO", "TRANSFERENCIA")}
    totais["SALDO_BLOQUEADO"] = round(
        totais["BLOQUEIO"] - totais["DESBLOQUEIO"] - totais["TRANSFERENCIA"], 2
    )

    saldo_conta_valor = round(float(saldo_conta), 2) if saldo_conta else 0.0
    garantia = apurar_garantia(totais, fluxos, saldo_conta_valor)

    datas = [l["data"] for l in lancamentos]
    extrato = {
        "arquivo": caminho.name,
        "titular": meta.get("titular", ""),
        "agencia": meta.get("agencia", ""),
        "conta": meta.get("conta", ""),
        "periodo": meta.get("periodo", ""),
        "gerado_em": meta.get("gerado_em", ""),
        "importado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "primeiro_lancamento": min(datas) if datas else "",
        "ultimo_lancamento": max(datas) if datas else "",
        "saldo_conta_corrente": saldo_conta_valor,
        "estornos_de_transferencia": estornos,
        "totais": totais,
        "garantia": garantia,
        "fluxos": sorted(fluxos, key=lambda f: (f["data"], f["tipo"])),
        "lancamentos": sorted(lancamentos, key=lambda l: (l["data"], l["tipo"])),
    }

    EXTRATO_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(EXTRATO_FILE, "w", encoding="utf-8") as f:
        json.dump(extrato, f, ensure_ascii=False, indent=2)

    print(f"Extrato importado: {caminho.name}")
    print(f"  Conta ............... {extrato['agencia']}/{extrato['conta']}")
    print(f"  Período ............. {extrato['primeiro_lancamento']} a {extrato['ultimo_lancamento']}")
    print(f"  Lançamentos ......... {len(lancamentos)} "
          f"({estornos} desbloqueios de estorno neutralizados)")
    print(f"  Bloqueado ........... {totais['BLOQUEIO']:,.2f}")
    print(f"  Desbloqueado ........ {totais['DESBLOQUEIO']:,.2f}")
    print(f"  Transferido ......... {totais['TRANSFERENCIA']:,.2f}")
    print(f"  Saldo bloqueado ..... {totais['SALDO_BLOQUEADO']:,.2f}")
    print(f"  Saldo em c/c ........ {extrato['saldo_conta_corrente']:,.2f}")
    print(f"  Créditos ............ {garantia['creditos']:,.2f}")
    print(f"  Rendimentos ......... {garantia['rendimentos']:,.2f}")
    print(f"\n--- SALDO MÍNIMO RETIDO ---")
    print(f"  Exigido ............. {garantia['saldo_minimo_exigido']:,.2f}")
    print(f"  Colchão efetivo ..... {garantia['colchao_efetivo']:,.2f} "
          f"({garantia['cobertura'] * 100:.2f}% do exigido)")
    print(f"  Déficit ............. {garantia['deficit']:,.2f}")
    print(f"  Fora da conta ....... {garantia['fora_da_conta']:,.2f} "
          f"(perto de zero indica que nao ha recursos aplicados)")
    print(f"\nSalvo em {EXTRATO_FILE}")
    return extrato


def neutralizar_estornos(brutos):
    """Remove, dia a dia, um DESBLOQUEIO para cada TRANSFERENCIA de mesmo valor."""
    por_dia = {}
    for l in brutos:
        por_dia.setdefault(l["data"], {"BLOQUEIO": [], "DESBLOQUEIO": [], "TRANSFERENCIA": []})
        por_dia[l["data"]][l["tipo"]].append(l["valor"])

    lancamentos, estornos = [], 0
    for dia in sorted(por_dia):
        desbloqueios = list(por_dia[dia]["DESBLOQUEIO"])
        for valor in por_dia[dia]["BLOQUEIO"]:
            lancamentos.append({"data": dia, "tipo": "BLOQUEIO", "valor": valor})
        for valor in por_dia[dia]["TRANSFERENCIA"]:
            lancamentos.append({"data": dia, "tipo": "TRANSFERENCIA", "valor": valor})
            if valor in desbloqueios:
                desbloqueios.remove(valor)
                estornos += 1
        for valor in desbloqueios:
            lancamentos.append({"data": dia, "tipo": "DESBLOQUEIO", "valor": valor})
    return lancamentos, estornos


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    importar(sys.argv[1])
